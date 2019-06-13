import requests,json,unittest,os,sys
import decimal
from openpyxl import load_workbook
#path_data=os.path.join(os.path.split(os.getcwd())[0],"testdata")
path_data=os.path.join(os.getcwd(),"testdata")
path_report=os.path.join(os.getcwd(),"common")
#path1=os.path.join(os.path.split(os.getcwd())[0],"common")
sys.path.append("../")
sys.path.append(path_report)
#from exclepase import *
from config import *
from report_write import *
from logger import *
from sql_con import *
file_name="搜索商品.xlsx"
url_login = base_url+'zlead/member/login'
url = base_url+'query/goods'
mylogger = Logger(__name__).getlog()
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path_data,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.keys = list(self.sheet["B"])
        self.addressId = list(self.sheet["C"])
        self.ex_res= list(self.sheet["D"])
        self.re_res = list(self.sheet["E"])
        self.is_pass = list(self.sheet["F"])
        self.nrows=self.sheet.max_row
    
    def test_查询购物车(self):
        #登录
        pass_count = 0
        fail_count = 0
        key=self.keys[3].value
        seesion = requests.session()
        mylogger.info("登录开始")
        data = {'account':str(self.sheet["B2"].value),'password':str(self.sheet["C2"].value)}
        #res = requests.get(url=url_login,params=data)
        res = seesion.get(url=url_login,params = data)
        cookies=res.cookies
        if "登录成功" in res.json().values():
            mylogger.info("登录成功")
        else:
            mylogger.info("登录失败,失败信息为："+res.json())
            fail_count = 1
        mylogger.info("搜索商品开始，搜索关键字："+key)
        
        headers = {"Cookie":cookies}
        data_query = {'sour':3,"key":key,'f':'','b':'','m':'','l':'','c':'','brandids':'','listIds':'','catids':'','x':'','x':'','s':10}
        tt=seesion.get(url=url,params=data_query)
        #print(tt.json())
        mylogger.info("请求数据:"+str(tt.json()))
        n = len(tt.json()["data"])

        sum = ((tt.json()["data"])['count'])
        print("搜索名称中包含"+key+"的商品数量："+str(sum))
        print("执行SQL查询",end='')
        sql="SELECT COUNT(*) from t_goods   where full_name LIKE '%"+key+"%'"
        mylogger.info("执行SQL查询:"+str(sql))
        ttt =list(fetchone(sql))[0]
        mylogger.info("数据库查询购物车数量："+str(ttt))
        if str(sum)==str(ttt):
            pass_count=1
            fail_count=0
            mylogger.info("搜索商品成功")
        else:
            fail_count=1
            mylogger.info("搜索商品失败")
        content = ('''<tr><td>搜索商品</td><td>'''+str(pass_count+fail_count)+'''</td><td>'''+str(pass_count)+'''</td><td>'''+str(fail_count)+'''</td></tr>''')
        mylogger.info("搜索商品接口测试结果:用例总数"+str(pass_count+fail_count)+"条，其中通过用例"+str(pass_count)+"条，失败用例"+str(fail_count)+'条')
        #print(content)
        report(content).write_file()            
#if __name__=='__main__':
test_login().test_查询购物车()
