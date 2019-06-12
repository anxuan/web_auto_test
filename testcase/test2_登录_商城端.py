import requests,json,unittest,os,sys
from openpyxl import load_workbook
path=os.path.join(os.getcwd(),"testdata")
#path_report
path_report=os.path.join(os.getcwd(),"common")
path_log=os.path.join(os.getcwd(),"logs")
#print("path_report"+path_report)
sys.path.append(path_report)
sys.path.append("../")
#print(sys.path)
#from exclepase import *
from logger import *
from config import *
from report_write import *
from read_log import *
mylogger = Logger(__name__).getlog()
file_name="登录_商城端.xlsx"
url = base_url+'zlead/member/login'
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
print("*****************"+new_log(path_log))
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.username = list(self.sheet["B"])
        self.pwd = list(self.sheet["C"])
        self.ex_res= list(self.sheet["D"])
        self.re_res = list(self.sheet["E"])
        self.is_pass = list(self.sheet["F"])
        self.nrows=self.sheet.max_row
    
    def test_登录(self):
        pass_count = 0
        fail_count = 0
        for i in range(1,self.nrows):
            mylogger.info("第"+str(i)+"条用例\""+str(self.description[i].value)+"\"开始执行")
            data={'account':str(self.username[i].value),'password':str(self.pwd[i].value)}
            mylogger.info("请求数据:"+str(data))
            p=requests.get(url=url,params=data)
            mylogger.info("返回数据:"+str(p.json()))
            self.sheet[('E'+str(i+1))]=str(p.json())
            self.data.save(self.real_path)
            s=requests.session()
            if self.ex_res[i].value in p.json().values():
                self.sheet[('F'+str(i+1))]="成功"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i)+"条用例执行结束，执行结果:通过!\n")
                pass_count=pass_count+1
            else:
                self.sheet[('F'+str(i+1))]="失败"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i)+"条用例执行结束，执行结果:失败")
                fail_count=fail_count+1
        content = ('''<tr><td>登录</td><td>'''+str(pass_count+fail_count)+'''</td><td>'''+str(pass_count)+'''</td><td>'''+str(fail_count)+'''</td></tr>''')
        mylogger.info("登录接口测试结果:用例总数"+str(pass_count+fail_count)+"条，其中通过用例"+str(pass_count)+"条，失败用例"+str(fail_count)+'条')
        report(content).write_file()
            
#if __name__=='__main__':
test_login().test_登录()
