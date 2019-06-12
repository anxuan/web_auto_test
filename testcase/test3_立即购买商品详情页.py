import requests,json,unittest,os,sys
import decimal
from openpyxl import load_workbook
#path_data=os.path.join(os.path.split(os.getcwd())[0],"testdata")
path_data=os.path.join(os.getcwd(),"testdata")
path_report=os.path.join(os.getcwd(),"common")
#path1=os.path.join(os.path.split(os.getcwd())[0],"common")
sys.path.append("../")
sys.path.append(path_report)
#from exclepase import *
from config import *
from report_write import *
from logger import *
from sql_con import *
file_name="立即购买(商品详情).xlsx"
url_login = base_url+'zlead/member/login'
url = base_url+'/zlead/order/newConfirmOrder'
mylogger = Logger(__name__).getlog()
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path_data,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.goodsId = list(self.sheet["B"])
        self.buyNum = list(self.sheet["C"])
        self.ex_res= list(self.sheet["D"])
        self.re_res = list(self.sheet["E"])
        self.is_pass = list(self.sheet["F"])
        self.nrows=self.sheet.max_row
    
    def test_立即购买(self):
        #登录
        pass_count = 0
        fail_count = 0
        print("********************************************************************")
        seesion = requests.session()
        data = {'account':str(self.sheet["B2"].value),'password':str(self.sheet["C2"].value)}
        #res = requests.get(url=url_login,params=data)
        res = seesion.get(url=url_login,params = data)
        cookies=res.cookies
        print(self.nrows)
        for i in range(3,self.nrows):
            mylogger.info("第"+str(i-2)+"条用例\""+str(self.description[i].value)+"\"开始执行")
            headers = {"Cookie":cookies}
            data = {'goodsId':self.goodsId[i].value,'buyNum':self.buyNum[i].value}
            mylogger.info("请求数据:"+str(data))
            p=seesion.post(url=url,data=data,cookies=cookies)
            #print(tt.json())
            self.sheet[('E'+str(i+1))]=str(p.json())            
            self.data.save(self.real_path)
            mylogger.info("返回数据:"+str(p.json()))
            n = len(p.json()["data"])
            sum=0
            if self.ex_res[i].value in p.json().values():
                self.sheet[('F'+str(i+1))]="成功"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i-2)+"条用例执行结束，执行结果:通过!\n")
                pass_count=pass_count+1
                print(pass_count)
            else:
                self.sheet[('F'+str(i+1))]="失败"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i-2)+"条用例执行结束，执行结果:失败")
                fail_count=fail_count+1
        content = ('''<tr><td>立即购买(商品详情)</td><td>'''+str(pass_count+fail_count)+'''</td><td>'''+str(pass_count)+'''</td><td>'''+str(fail_count)+'''</td></tr>''')
        mylogger.info("立即购买(商品详情)接口测试结果:用例总数"+str(pass_count+fail_count)+"条，其中通过用例"+str(pass_count)+"条，失败用例"+str(fail_count)+'条')
        #print(content)
        report(content).write_file()          
#if __name__=='__main__':
test_login().test_立即购买()
