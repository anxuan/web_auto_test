import requests,json,unittest,os,sys
from openpyxl import load_workbook
#path=os.path.join(os.path.split(os.getcwd())[0],"testdata")
path_data=os.path.join(os.getcwd(),"testdata")
path_report=os.path.join(os.getcwd(),"common")
sys.path.append(path_report)
sys.path.append("../")
#from exclepase import *
from report_write import *
from config import *
from report_write import *
from logger import *
file_name="注册.xlsx"
url = base_url+'zlead/member/registered'
mylogger = Logger(__name__).getlog()
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path_data,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.username = list(self.sheet["B"])
        self.pwd = list(self.sheet["C"])
        self.company = list(self.sheet["D"])
        self.vcode = list(self.sheet["E"])
        self.ex_res= list(self.sheet["F"])
        self.re_res = list(self.sheet["G"])
        self.is_pass = list(self.sheet["H"])
        self.nrows=self.sheet.max_row
    
    def test_login(self):
        pass_count = []
        fail_count = []
        for i in range(1,self.nrows):
            mylogger.info("第"+str(i)+"条用例\""+str(self.description[i].value)+"\"开始执行")
            #data = {'phone': '13141255148', 'passWord': '111111aa', 'companyName': 'test', 'verifyCode': '111111'}
            data={'phone':str(self.username[i].value),'passWord':str(self.pwd[i].value),'companyName':str(self.company[i].value),'verifyCode':str(self.vcode[i].value)}
            mylogger.info("请求数据:"+str(data))
            p=requests.post(url,data=data)
            #print("返回数据:"+str(p.json()))
            mylogger.info("返回数据:"+str(p.json()))
            #mylogger.info(p.json())
            self.sheet[('G'+str(i+1))]=str(p.json())            
            self.data.save(self.real_path)
            s=requests.session()
            if self.ex_res[i].value in p.json().values():
                self.sheet[('H'+str(i+1))]="成功"
                self.data.save(self.real_path)
                pass_count.append(i)
                mylogger.info("第"+str(i)+"条用例执行结束，执行结果:通过!\n")
            else:
                self.sheet[('H'+str(i+1))]="失败"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i)+"条用例执行结束，执行结果:失败")
                fail_count.append(i)
        content = ('''<h1 align = "center">直链网F端接口自动化测试报告</h1><table border = "3"><tr><th>模块</th><th>用例总数</th><th>用例通过数</th><th>用例失败数</th></tr><tr><td>注册</td><td>'''+str(self.nrows-1)+'''</td><td>'''+str(len(pass_count))+'''</td><td>'''+str(len(fail_count))+'''</td></tr>''')
        mylogger.info("登录接口测试结果:用例总数"+str(pass_count+fail_count)+"条，其中通过用例"+str(pass_count)+"条，失败用例"+str(fail_count)+'条') 
        #content=("注册模块测试结果:用例总数"+str(self.nrows-1)+"其中通过用例"+str(len(pass_count))+"条，失败用例"+str(len(fail_count))+'条')
        report(content).write_file()            
#if __name__=='__main__':
test_login().test_login()
