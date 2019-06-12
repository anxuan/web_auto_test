import requests,json,unittest,os,sys
import decimal
from openpyxl import load_workbook
#path_data=os.path.join(os.path.split(os.getcwd())[0],"testdata")
path_data=os.path.join(os.getcwd(),"testdata")
path_report=os.path.join(os.getcwd(),"common")
#path1=os.path.join(os.path.split(os.getcwd())[0],"common")
sys.path.append("../")
sys.path.append(path_report)
#from exclepase import *
from config import *
from report_write import *
from logger import *
from sql_con import *
file_name="提交订单.xlsx"
url_login = base_url+'zlead/member/login'
url = base_url+'/zlead/order/addOrder'
mylogger = Logger(__name__).getlog()
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path_data,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.buyNum = list(self.sheet["B"])
        self.goodsId = list(self.sheet["C"])
        self.addressId = list(self.sheet["D"])
        self.buyType = list(self.sheet["E"])
        self.storeBuyType = list(self.sheet["F"])
        self.orderType = list(self.sheet["G"])
        self.ex_res= list(self.sheet["H"])
        self.re_res = list(self.sheet["I"])
        self.is_pass = list(self.sheet["J"])
        self.nrows=self.sheet.max_row
    
    def test_提交订单(self):
        #登录
        pass_count = 0
        fail_count = 0
        seesion = requests.session()
        data = {'account':str(self.sheet["B2"].value),'password':str(self.sheet["C2"].value)}
        #res = requests.get(url=url_login,params=data)
        res = seesion.get(url=url_login,params = data)
        cookies=res.cookies
        for i in range(3,self.nrows):
            mylogger.info("第"+str(i-2)+"条用例\""+str(self.description[i].value)+"\"开始执行")
            data={'goodsId':str(self.goodsId[i].value),'buyNum':str(self.buyNum[i].value),'buyType':str(self.buyType[i].value),'addressId':str(self.addressId[i].value),'storeBuyType':str(self.storeBuyType[i].value),'orderType':str(self.orderType[i].value)}
            headers = {"Cookie":cookies}
            mylogger.info("请求数据:"+str(data))
            p=seesion.post(url=url,data=data,cookies=cookies)
            mylogger.info("返回数据:"+str(p.json()))
            self.sheet[('I'+str(i+1))]=str(p.json())            
            self.data.save(self.real_path)
            mylogger.info("预期结果："+self.ex_res[i].value)
            if self.ex_res[i].value in p.json().values():
                self.sheet[('J'+str(i+1))]="成功"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i-2)+"条用例执行结束，执行结果:通过!\n")
                pass_count=+1
                print(pass_count)
                
            else:
                self.sheet[('J'+str(i+1))]="失败"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i-2)+"条用例执行结束，执行结果:失败")
                fail_count=+1    
        content = ('''<tr><td>提交订单</td><td>'''+str(pass_count+fail_count)+'''</td><td>'''+str(pass_count)+'''</td><td>'''+str(fail_count)+'''</td></tr>''')
        mylogger.info("提交订单接口测试结果:用例总数"+str(pass_count+fail_count)+"条，其中通过用例"+str(pass_count)+"条，失败用例"+str(fail_count)+'条')
        #print(content)
        report(content).write_file()            
#if __name__=='__main__':
test_login().test_提交订单()
