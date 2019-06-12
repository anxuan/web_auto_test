import requests,json,unittest,os,sys
import decimal
from openpyxl import load_workbook
#path_data=os.path.join(os.path.split(os.getcwd())[0],"testdata")
path_data=os.path.join(os.getcwd(),"testdata")
path_report=os.path.join(os.getcwd(),"common")
#path1=os.path.join(os.path.split(os.getcwd())[0],"common")
sys.path.append("../")
sys.path.append(path_report)
#from exclepase import *
from config import *
from report_write import *
from logger import *
from sql_con import *
file_name="查看购物车.xlsx"
url_login = base_url+'zlead/member/login'
url = base_url+'zlead/shopcart/shoppingCartGoods'
mylogger = Logger(__name__).getlog()
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path_data,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.shopIds = list(self.sheet["B"])
        self.addressId = list(self.sheet["C"])
        self.cartIds = list(self.sheet["D"])
        self.buyType = list(self.sheet["E"])
        self.storeBuyType = list(self.sheet["F"])
        self.orderType = list(self.sheet["G"])
        self.ex_res= list(self.sheet["H"])
        self.re_res = list(self.sheet["I"])
        self.is_pass = list(self.sheet["J"])
        self.nrows=self.sheet.max_row
    
    def test_查询购物车(self):
        #登录
        pass_count = 0
        fail_count = 0
        seesion = requests.session()
        data = {'account':str(self.sheet["B2"].value),'password':str(self.sheet["C2"].value)}
        #res = requests.get(url=url_login,params=data)
        res = seesion.get(url=url_login,params = data)
        cookies=res.cookies
        
        headers = {"Cookie":cookies}
        tt=seesion.get(url)
        #print(tt.json())
        mylogger.info("请求数据:"+str(tt.json()))
        n = len(tt.json()["data"])
        sum=0
            
        for i in range(n):
            sum += ((tt.json()["data"])[i]['goodsList'][0]['count'])
            print((tt.json()["data"])[i]['goodsList'][0]['count'])
        print("接口查询购物车数量："+str(sum))
        print("执行SQL查询",end='')
        sql="SELECT sum(count) from t_shopping_cart where member_id = (SELECT id from t_member where member_id = \'"+str(self.sheet["B2"].value)+"\') and is_buy = \'0\'"
        mylogger.info("执行SQL查询:"+str(sql))
        ttt =list(fetchone(sql))[0]
        print("数据库查询购物车数量："+str(ttt))
        if str(sum)==str(ttt):
            pass_count=+1
            mylogger.info("查询购物车成功")
        else:
            fail_count=+1
            mylogger.info("查询购物车错误")
        content = ('''<tr><td>查看购物车</td><td>'''+str(pass_count+fail_count)+'''</td><td>'''+str(pass_count)+'''</td><td>'''+str(fail_count)+'''</td></tr>''')
        mylogger.info("查看购物车接口测试结果:用例总数"+str(pass_count+fail_count)+"条，其中通过用例"+str(pass_count)+"条，失败用例"+str(fail_count)+'条')
        #print(content)
        report(content).write_file()            
#if __name__=='__main__':
test_login().test_查询购物车()
