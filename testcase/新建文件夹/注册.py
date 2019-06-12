import requests,json,unittest,os,sys
from openpyxl import load_workbook
path=os.path.join(os.path.split(os.getcwd())[0],"testdata")

sys.path.append("../")
#from exclepase import *
from config import *
file_name="注册.xlsx"
url = base_url+'zlead/member/registered'
print("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.username = list(self.sheet["B"])
        self.pwd = list(self.sheet["C"])
        self.company = list(self.sheet["D"])
        self.vcode = list(self.sheet["E"])
        self.ex_res= list(self.sheet["F"])
        self.re_res = list(self.sheet["G"])
        self.is_pass = list(self.sheet["H"])
        self.nrows=self.sheet.max_row
    
    def test_login(self):
        for i in range(1,self.nrows):
            print("第"+str(i)+"条用例\""+str(self.description[i].value)+"\"开始执行")
            data = {'phone': '13141255148', 'passWord': '111111aa', 'companyName': 'test', 'verifyCode': '111111'}
            #data={'phone':str(self.username[i].value),'passWord':str(self.pwd[i].value),'companyName':str(self.company[i].value),'verifyCode':str(self.vcode[i].value)}
            print("请求数据:"+str(data))
            p=requests.post(url,data=data)
            #print("返回数据:"+str(p.json()))
            print("返回数据:",end='')
            print(p.json())
            self.sheet[('G'+str(i+1))]=str(p.json())            
            self.data.save(self.real_path)
            s=requests.session()
            if self.ex_res[i].value in p.json().values():
                self.sheet[('H'+str(i+1))]="成功"
                self.data.save(self.real_path)
                
                print("第"+str(i)+"条用例执行结束，执行结果:通过!\n")
            else:
                self.sheet[('H'+str(i+1))]="失败"
                self.data.save(self.real_path)
                print("第"+str(i)+"条用例执行结束，执行结果:失败")
            
if __name__=='__main__':
    test_login().test_login()
