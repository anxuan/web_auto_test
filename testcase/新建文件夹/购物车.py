import requests,json,unittest,os,sys
import decimal
from openpyxl import load_workbook
path=os.path.join(os.path.split(os.getcwd())[0],"testdata")
path1=os.path.join(os.path.split(os.getcwd())[0],"common")
sys.path.append("../")
sys.path.append(path1)
#from exclepase import *
from config import *
from logger import *
from sql_con import *
file_name="订单.xlsx"
url_login = base_url+'zlead/member/login'
url_look = base_url+'zlead/shopcart/shoppingCartGoods'
url = base_url + 'zlead/shopcart/addShoppingCart'
#print("本次接口测试请求地址为："+str(url)+"\n")
mylogger = Logger(logger='TestMyLog').getlog()
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.goodsId = list(self.sheet["B"])
        self.count = list(self.sheet["C"])
        self.buytype = list(self.sheet["D"])
        self.ex_res= list(self.sheet["E"])
        self.re_res = list(self.sheet["F"])
        self.is_pass = list(self.sheet["G"])
        self.nrows=self.sheet.max_row
    
    def shopping_cart(self):
        #登录
        seesion = requests.session()
        data = {'account':str(self.sheet["B2"].value),'password':str(self.sheet["C2"].value)}
        #res = requests.get(url=url_login,params=data)
        res = seesion.get(url=url_login,params = data)
        cookies=res.cookies
        mylogger.info("获取cookie成功，cookies="+str(cookies))
        if "登录成功" in res.json().values():
            mylogger.info("登录成功")
        else:
            mylogger.info("登录失败"+res.json().values()[1])
        print(self.nrows)
        for i in range(3,self.nrows):
            mylogger.info("第"+str(i-2)+"条用例\""+str(self.description[i].value)+"\"开始执行")
            #print(self.username[i].value,self.pwd[i].value)
            data={'goodsId':str(self.goodsId[i].value),'count':str(self.count[i].value),'buyType':str(self.buytype[i].value)}
            mylogger.info("请求数据:"+str(data))
            #p=requests.post(url=url,data=data,cookies=cookies)
            p=seesion.post(url=url,data=data,cookies=cookies)
            #print("返回数据:"+str(p.json()))
            mylogger.info("返回数据:"+str(p.json()))
            #print(list(p.json()).count())
            self.sheet[('F'+str(i+1))]=str(p.json())            
            self.data.save(self.real_path)
            mylogger.info(self.ex_res[i].value)
            if self.ex_res[i].value in p.json().values():
                self.sheet[('G'+str(i+1))]="成功"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i-2)+"条用例执行结束，执行结果:通过!\n")
            else:
                self.sheet[('G'+str(i+1))]="失败"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i-2)+"条用例执行结束，执行结果:失败")
            print("查询购物车开始")
            headers = {"Cookie":cookies}
            tt=seesion.get(url_look)
            n = len(tt.json()["data"])
            sum=0
            
            for i in range(n):
                sum += ((tt.json()["data"])[i]['goodsList'][0]['count'])
                print((tt.json()["data"])[i]['goodsList'][0]['count'])
            print("接口查询购物车数量："+str(sum))
            print("执行SQL查询",end='')
            sql="SELECT sum(count) from t_shopping_cart where member_id = (SELECT id from t_member where member_id = \'"+str(self.sheet["B2"].value)+"\') and is_buy = \'0\'"
            #print(sql)
            ttt =list(fetchone(sql))[0]
            print("数据库查询购物车数量："+str(ttt))
            
if __name__=='__main__':
    test_login().shopping_cart()
