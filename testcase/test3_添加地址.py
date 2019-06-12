import requests,json,unittest,os,sys
import decimal
from openpyxl import load_workbook
#path_data=os.path.join(os.path.split(os.getcwd())[0],"testdata")
path_data=os.path.join(os.getcwd(),"testdata")
path_report=os.path.join(os.getcwd(),"common")
#path1=os.path.join(os.path.split(os.getcwd())[0],"common")
sys.path.append("../")
sys.path.append(path_report)
#from exclepase import *
from config import *
from report_write import *
from logger import *
from sql_con import *
from md5parse import *
file_name="添加地址.xlsx"
url_login = base_url+'zlead/member/login'
url_look = base_url+'zlead/shopcart/shoppingCartGoods'
url = base_url + 'zlead/memaddr/addOrUpdateMemberAddress'
#print("本次接口测试请求地址为："+str(url)+"\n")
mylogger = Logger(__name__).getlog()
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path_data,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.memberName = list(self.sheet["B"])
        self.phone = list(self.sheet["C"])
        self.provinceName = list(self.sheet["D"])
        self.cityName = list(self.sheet["E"])
        self.countyName = list(self.sheet["F"])
        self.detailAddress = list(self.sheet["G"])
        self.isDefault = list(self.sheet["H"])
        self.type = list(self.sheet["I"])
        self.ex_res= list(self.sheet["J"])
        self.re_res = list(self.sheet["K"])
        self.is_pass = list(self.sheet["L"])
        self.nrows=self.sheet.max_row
    
    def test_添加地址(self):
        #登录
        pass_count = []
        fail_count = []
        seesion = requests.session()
        data = {'account':str(self.sheet["B2"].value),'password':str(self.sheet["C2"].value)}
        #res = requests.get(url=url_login,params=data)
        res = seesion.get(url=url_login,params = data)
        cookies=res.cookies
        mylogger.info("获取cookie成功，cookies="+str(cookies))
        if "登录成功" in res.json().values():
            mylogger.info("登录成功")
        else:
            mylogger.info("登录失败,失败信息为："+res.json())
        print(self.nrows)
        for i in range(3,self.nrows):
            mylogger.info("第"+str(i-2)+"条用例\""+str(self.description[i].value)+"\"开始执行")
            #print(self.username[i].value,self.pwd[i].value)
            #data={'memberName':md5parse(str(self.memberName[i].value).encode()),'phone':str(self.phone[i].value),'provinceName':md5parse(str(self.provinceName[i].value).encode()),'cityName':md5parse(str(self.cityName[i].value).encode()),'countyName':md5parse(str(self.countyName[i].value).encode()),'detailAddress':md5parse(str(self.detailAddress[i].value).encode()),'isDefault':str(self.isDefault[i].value),'type':str(self.type[i].value)}
            #print(int(self.count[i].value))
            data = {'memberName':str(self.memberName[i].value),'phone':str(self.phone[i].value),'provinceName':str(self.provinceName[i].value),'cityName':str(self.cityName[i].value),'countyName':str(self.countyName[i].value),'detailAddress':str(self.detailAddress[i].value),'isDefault':str(self.isDefault[i].value),'type':str(self.type[i].value)}
            mylogger.info("请求数据:"+str(data))
            #p=requests.post(url=url,data=data,cookies=cookies)
            p=seesion.post(url=url,data=data,cookies=cookies)
            #print("返回数据:"+str(p.json()))
            mylogger.info("返回数据:"+str(p.text))
            #print(list(p.json()).count())
            
            self.sheet[('K'+str(i+1))]=str(p.json())            
            self.data.save(self.real_path)
            mylogger.info(self.ex_res[i].value)
           
            if self.ex_res[i].value in p.json().values():
                self.sheet[('L'+str(i+1))]="成功"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i-2)+"条用例执行结束，执行结果:通过!\n")
                #self.sheet[('F'+str(i+1))]="失败"
                pass_count.append(i)
                
            else:
                self.sheet[('L'+str(i+1))]="失败"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i-2)+"条用例执行结束，执行结果:失败")
                fail_count.append(i)
                print(fail_count)

                    
        content = ('''<tr><td>添加地址</td><td>'''+str(self.nrows-3)+'''</td><td>'''+str(len(pass_count))+'''</td><td>'''+str(len(fail_count))+'''</td></tr>''')
        mylogger.info('''添加地址模块测试结果:用例总数'''+str(self.nrows-3)+"其中通过用例"+str(len(pass_count))+"条，失败用例"+str(len(fail_count))+'条')
        print(content)
        report(content).write_file()

#if __name__=='__main__':
test_login().test_添加地址()
