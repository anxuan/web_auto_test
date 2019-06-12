import requests,json,unittest,os,sys
import decimal
from openpyxl import load_workbook
#path_data=os.path.join(os.path.split(os.getcwd())[0],"testdata")
path_data=os.path.join(os.getcwd(),"testdata")
path_report=os.path.join(os.getcwd(),"common")
#path1=os.path.join(os.path.split(os.getcwd())[0],"common")
sys.path.append("../")
sys.path.append(path_report)
#from exclepase import *
from config import *
from report_write import *
from logger import *
from sql_con import *
file_name="发送验证码.xlsx"
url = base_url+'zlead/member/sendVerifyCode'
mylogger = Logger(__name__).getlog()
mylogger.info("本次接口测试请求地址为："+str(url)+"\n")
class test_login:
    def __init__(self,file_name=file_name,Sheet_name='test'):
        self.name=file_name
        self.Sheet_name = Sheet_name
        self.real_path=os.path.join(path_data,self.name)
        self.data = load_workbook(self.real_path)        
        self.sheet = self.data[self.Sheet_name]
        self.description = list(self.sheet["A"])
        self.phone = list(self.sheet["B"])
        self.ex_res= list(self.sheet["C"])
        self.re_res = list(self.sheet["D"])
        self.is_pass = list(self.sheet["E"])
        self.nrows=self.sheet.max_row
    
    def test_发送验证码(self):
        #登录
        pass_count = 0
        fail_count = 0
        seesion = requests.session()
        for i in range(1,self.nrows):
            mylogger.info("第"+str(i)+"条用例\""+self.description[i].value+"\"执行开始")
            data = {'phone':str(self.phone[i].value)}
            mylogger.info("发出请求数据："+url+str(data))
            res = requests.post(url =url,data = data)
            mylogger.info("返回请求数据："+str(res.json()))
            if self.ex_res[i].value in res.json().values():
                self.re_res[i]="成功"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i)+"条用例执行结束，执行结果:通过!\n")
                pass_count=+1
            else:
                self.re_res[i]="失败"
                self.data.save(self.real_path)
                mylogger.info("第"+str(i)+"条用例执行结束，执行结果:失败")
                fail_count=+1
        content = ('''<tr><td>发送验证码</td><td>'''+str(pass_count+fail_count)+'''</td><td>'''+str(pass_count)+'''</td><td>'''+str(fail_count)+'''</td></tr>''')
        mylogger.info("发送验证码接口测试结果:用例总数"+str(pass_count+fail_count)+"条，其中通过用例"+str(pass_count)+"条，失败用例"+str(fail_count)+'条')
        report(content).write_file()
#if __name__=='__main__':
test_login().test_发送验证码()
